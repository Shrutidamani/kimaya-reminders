import requests
import openpyxl
import io
import re
from datetime import datetime, timedelta

class SheetsClient:
    def __init__(self, url):
        self.url = url
        self.spreadsheet_id, self.gid = self._parse_url(url)

    def _parse_url(self, url):
        """Extract Spreadsheet ID and GID from the Google Sheet URL."""
        spreadsheet_id = ""
        gid = "0"
        
        id_match = re.search(r"/d/([a-zA-Z0-9-_]{30,60})", url)
        if id_match:
            spreadsheet_id = id_match.group(1)
            
        gid_match = re.search(r"gid=(\d+)", url)
        if gid_match:
            gid = gid_match.group(1)
            
        return spreadsheet_id, gid

    def is_valid(self):
        return bool(self.spreadsheet_id)

    def _is_green(self, cell):
        """Check if a cell's background fill is green."""
        if not cell or not cell.fill:
            return False
        
        fill = cell.fill
        if fill.fill_type is None or fill.fill_type == 'none':
            return False
            
        color = fill.fgColor
        if not color:
            return False
            
        rgb = color.rgb
        if not rgb or not isinstance(rgb, str):
            return False
            
        hex_val = rgb.upper().strip()
        if len(hex_val) == 8:
            hex_val = hex_val[2:] # Strip Alpha channel
            
        if len(hex_val) != 6:
            return False
            
        try:
            r = int(hex_val[0:2], 16)
            g = int(hex_val[2:4], 16)
            b = int(hex_val[4:6], 16)
            
            common_greens = [
                "C6EFCE", "E2EFDA", "D4EDDA", "D5E8D4", "B7F4C7", "A9DFBF", "A3E4D7",
                "D1F2EB", "D5F5E3", "90EE90", "98FB98", "8FBC8F", "A2D9CE", "2ECC71",
                "27AE60", "00FF00", "00FF7F", "39FF14", "D0F0C0", "E2F0D9", "C6E0B4"
            ]
            
            if any(cg in hex_val for cg in common_greens):
                return True
                
            if g > r + 15 and g > b + 15 and g > 60:
                return True
        except Exception:
            pass
            
        return False

    def _parse_date(self, val):
        """Convert datetime object, string, or Excel date serial into standard YYYY-MM-DD string."""
        if val is None:
            return ""
            
        if isinstance(val, datetime):
            if val.year < 1901: # Filter out Excel default epoch dates
                return ""
            return val.strftime("%Y-%m-%d")
            
        if isinstance(val, (int, float)):
            if val == 0:
                return ""
            try:
                epoch = datetime(1899, 12, 30)
                res_date = epoch + timedelta(days=val)
                if res_date.year < 1901:
                    return ""
                return res_date.strftime("%Y-%m-%d")
            except Exception:
                return ""
                
        if isinstance(val, str):
            val_str = val.strip()
            if not val_str or val_str == "0":
                return ""
            # Try common date formats
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%b-%Y", "%d %b %Y"):
                try:
                    parsed = datetime.strptime(val_str, fmt)
                    if parsed.year < 1901:
                        return ""
                    return parsed.strftime("%Y-%m-%d")
                except ValueError:
                    continue
            return val_str
            
        return str(val)

    def fetch_records(self):
        """Download the spreadsheet and parse transactions from columns A, B, G, and Q."""
        if not self.is_valid():
            raise ValueError("Invalid Google Sheets URL")
            
        download_url = f"https://docs.google.com/spreadsheets/d/{self.spreadsheet_id}/export?format=xlsx&gid={self.gid}"
        
        response = requests.get(download_url, timeout=20)
        if response.status_code != 200:
            raise ConnectionError(
                f"Failed to download spreadsheet (Status Code: {response.status_code}). "
                "Ensure the spreadsheet is shared as 'Anyone with the link can view'."
            )
            
        wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)
        sheet = wb.active
        
        records = []
        
        # Columns map to: A=1 (Date), B=2 (Party), G=7 (Amount), Q=17 (Due Date), D=4 (Bill Number), AG=33 (Last Reminded), AH=34 (Reminder Count)
        for r in range(2, sheet.max_row + 1):
            date_cell = sheet.cell(row=r, column=1)
            party_cell = sheet.cell(row=r, column=2)
            amount_cell = sheet.cell(row=r, column=7)
            due_cell = sheet.cell(row=r, column=17) # Column Q
            bill_no_cell = sheet.cell(row=r, column=4) # Column D
            last_reminded_cell = sheet.cell(row=r, column=33) # Column AG
            reminder_count_cell = sheet.cell(row=r, column=34) # Column AH
            
            if party_cell.value is None and amount_cell.value is None:
                continue
                
            is_paid = self._is_green(party_cell) or self._is_green(amount_cell) or self._is_green(date_cell)
            
            raw_date = date_cell.value
            raw_due = due_cell.value
            
            formatted_date = self._parse_date(raw_date)
            formatted_due = self._parse_date(raw_due)
            
            party_name = str(party_cell.value).strip() if party_cell.value is not None else ""
            
            bill_no_val = ""
            if bill_no_cell.value is not None:
                b_val = bill_no_cell.value
                if isinstance(b_val, float) and b_val.is_integer():
                    bill_no_val = str(int(b_val))
                elif isinstance(b_val, datetime):
                    bill_no_val = b_val.strftime("%d-%m-%Y")
                else:
                    bill_no_val = str(b_val).strip()
            
            last_rem_val = ""
            if last_reminded_cell.value is not None:
                l_val = last_reminded_cell.value
                if isinstance(l_val, datetime):
                    last_rem_val = l_val.strftime("%Y-%m-%d %H:%M")
                else:
                    last_rem_val = str(l_val).strip()
                    
            rem_count_val = 0
            if reminder_count_cell.value is not None:
                r_val = reminder_count_cell.value
                try:
                    if isinstance(r_val, (int, float)):
                        rem_count_val = int(r_val)
                    elif isinstance(r_val, datetime):
                        # If cell was formatted as date in Excel, day of 1900 indicates serial integer
                        if r_val.year == 1900:
                            rem_count_val = r_val.day
                        else:
                            rem_count_val = 1
                    elif isinstance(r_val, str):
                        clean_str = "".join(c for c in r_val if c.isdigit())
                        if clean_str:
                            rem_count_val = int(clean_str)
                except Exception:
                    rem_count_val = 0
            
            amount_val = 0.0
            if amount_cell.value is not None:
                a_val = amount_cell.value
                if isinstance(a_val, (int, float)):
                    amount_val = float(a_val)
                else:
                    try:
                        amount_val = float(str(a_val).replace(",", "").replace("₹", "").replace("$", "").strip())
                    except Exception:
                        amount_val = 0.0
            
            bill_id = f"ROW-{r}"
            
            records.append({
                "bill_no": bill_id,
                "row_index": r,
                "date": formatted_date,
                "party": party_name,
                "amount": amount_val,
                "due_date": formatted_due,
                "status": "Paid" if is_paid else "Unpaid",
                "bill_number": bill_no_val,
                "last_reminded": last_rem_val,
                "reminder_count": rem_count_val
            })
            
        return records
